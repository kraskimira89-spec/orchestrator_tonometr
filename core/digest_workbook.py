"""
Excel-файл для вложения в письмо со сводкой уведомлений.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.daily_digest import STATUS_ORDER, STATUS_TITLE

STATUS_RU_ROW = {
    "green": "В норме",
    "yellow": "Скоро срок",
    "red": "Просрочено",
    "no_data": "Нет данных",
}


def build_digest_xlsx(digest: dict[str, Any], devices: list[dict]) -> bytes:
    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws = wb.active
    ws.title = "Сводка"
    ws.append(["Оркестратор Поверки — сводка уведомлений"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([f"Дата: {digest['date_str']}"])
    ws.append([])
    ws.append(["Показатель", "Количество"])
    for c in range(1, 3):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
    for st in STATUS_ORDER:
        ws.append([STATUS_TITLE[st], digest["by_status"].get(st, 0)])
    ws.append(["Всего приборов", digest["total"]])
    ws.append([])
    ws.append(["По срокам поверки", ""])
    for c in range(1, 3):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
    b = digest["buckets"]
    rows_b = [
        ("Просрочено", b["overdue"]),
        ("Осталось до 7 дней", b["within_7"]),
        ("Осталось 8–30 дней", b["within_30"]),
        ("Осталось 31–60 дней", b["within_60"]),
        ("Более 60 дней", b["beyond_60"]),
        ("Нет даты окончания", b["no_expiry"]),
    ]
    for label, n in rows_b:
        ws.append([label, n])
    ws.append([])
    ws.append(["Тип прибора", "Количество"])
    for c in range(1, 3):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
    for t, n in sorted(digest["by_type"].items(), key=lambda x: -x[1]):
        ws.append([t, n])
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14

    w2 = wb.create_sheet("По ответственным")
    w2.append(
        ["ФИО", "Всего", "Просрочено", "Скоро срок", "В норме", "Нет данных"]
    )
    for c in range(1, 7):
        cell = w2.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
    for fio, c in digest["by_fio"]:
        w2.append(
            [
                fio,
                c["total"],
                c["red"],
                c["yellow"],
                c["green"],
                c["no_data"],
            ]
        )
    w2.column_dimensions["A"].width = 36
    for col in "BCDEF":
        w2.column_dimensions[col].width = 12

    w3 = wb.create_sheet("Приборы")
    w3.append(
        [
            "ID",
            "Тип",
            "Инв. №",
            "Место",
            "Дата окончания",
            "Статус",
            "Ответственный",
        ]
    )
    for c in range(1, 8):
        cell = w3.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
    for d in devices:
        st = d.get("status") or "no_data"
        w3.append(
            [
                d.get("id"),
                d.get("type") or "",
                d.get("inventory_number") or "",
                d.get("location") or "",
                d.get("expiry_date") or "",
                STATUS_RU_ROW.get(st, st),
                d.get("responsible_fio") or "",
            ]
        )
    w3.column_dimensions["A"].width = 6
    w3.column_dimensions["B"].width = 12
    w3.column_dimensions["C"].width = 14
    w3.column_dimensions["D"].width = 48
    w3.column_dimensions["E"].width = 14
    w3.column_dimensions["F"].width = 14
    w3.column_dimensions["G"].width = 28
    for row in w3.iter_rows(min_row=2, max_row=w3.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = wrap

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
