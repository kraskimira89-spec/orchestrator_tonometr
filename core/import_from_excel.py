import os
import sys
from datetime import datetime
from typing import Optional

from openpyxl import load_workbook

# Добавляем корень проекта в sys.path, чтобы можно было импортировать db.database
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)

if PROJECT_ROOT not in sys.path:
    sys.path.append(PROJECT_ROOT)

from db.database import get_connection, init_database  # noqa: E402

BASE_DIR = PROJECT_ROOT
DATA_DIR = os.path.join(BASE_DIR, "data")
EXCEL_PATH = os.path.join(DATA_DIR, "journal.xlsx")

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
EXCEL_PATH = os.path.join(DATA_DIR, "journal.xlsx")


def parse_excel_date(value) -> Optional[str]:
    """Преобразует даты из Excel в строку YYYY-MM-DD или None."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str):
        try:
            # пробуем стандартные форматы
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    return datetime.strptime(value.strip(), fmt).date().isoformat()
                except ValueError:
                    continue
        except Exception:
            return None
    return None


def import_journal():
    """Импортирует приборы и поверки из Excel-файла journal.xlsx."""
    if not os.path.exists(EXCEL_PATH):
        print(f"Файл Excel не найден: {EXCEL_PATH}")
        return

    init_database()
    wb = load_workbook(EXCEL_PATH, data_only=True)

    conn = get_connection()
    cursor = conn.cursor()

    sheets_to_import = ["Алкометры", "Тонометры"]

    created_devices = 0
    created_verifications = 0

    for sheet_name in sheets_to_import:
        if sheet_name not in wb.sheetnames:
            print(f"⚠ Лист '{sheet_name}' не найден в файле, пропускаем.")
            continue

        sheet = wb[sheet_name]
        print(f"Импорт с листа: {sheet_name}")

        # предполагаем, что первая строка — заголовки
        header_row = 1

        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            dev_type = sheet_name  # тип берём из имени листа
            name = "" if row[0] is None else str(row[0]).strip()
            inventory_number = "" if row[1] is None else str(row[1]).strip()
            serial_number = "" if row[2] is None else str(row[2]).strip()
            location = "" if row[3] is None else str(row[3]).strip()
            responsible_name = "" if row[4] is None else str(row[4]).strip()
            verification_date = parse_excel_date(row[5])
            expiry_date = parse_excel_date(row[6])

            if not name:
                continue  # пустая строка

            # создаём/находим ответственного
            responsible_id = None
            if responsible_name:
                cursor.execute(
                    "SELECT id FROM users WHERE name = ?",
                    (responsible_name,)
                )
                res = cursor.fetchone()
                if res:
                    responsible_id = res["id"]
                else:
                    cursor.execute(
                        "INSERT INTO users (name, role) VALUES (?, 'responsible')",
                        (responsible_name,)
                    )
                    responsible_id = cursor.lastrowid

            # ищем, нет ли уже такого прибора
            cursor.execute(
                """
                SELECT id FROM devices
                WHERE name = ? AND IFNULL(inventory_number, '') = ? AND IFNULL(serial_number, '') = ?
                """,
                (name, inventory_number, serial_number),
            )
            dev = cursor.fetchone()
            if dev:
                device_id = dev["id"]
            else:
                cursor.execute(
                    """
                    INSERT INTO devices (type, name, inventory_number, serial_number, location, responsible_id)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (dev_type, name, inventory_number, serial_number, location, responsible_id),
                )
                device_id = cursor.lastrowid
                created_devices += 1

            # добавляем запись о поверке, если есть дата окончания
            if expiry_date:
                cursor.execute(
                    """
                    INSERT INTO verifications (device_id, verification_date, expiry_date, result, comment)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (device_id, verification_date, expiry_date, "пройдено", f"Импорт из {sheet_name}"),
                )
                created_verifications += 1

    conn.commit()
    conn.close()

    print(f"Импорт завершён.")
    print(f"Создано новых приборов: {created_devices}")
    print(f"Создано записей поверок: {created_verifications}")


if __name__ == "__main__":
    import_journal()
