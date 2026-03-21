"""
Умный импорт приборов из Excel с разрешением конфликтов (листы «Алкометры», «Тонометры»).
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from PyQt6.QtWidgets import (
    QComboBox,
    QDialog,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QFont

import openpyxl
from datetime import datetime, date

from db.database import get_connection

SHEETS = ["Алкометры", "Тонометры"]

SKIP_INV = {
    "",
    "инв. номер",
    "№",
    "инв.номер",
    "инвентарный номер",
    "итого",
    "итого:",
    "местонахождение (склад/вагон)",
}


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def load_excel_devices(path):
    """Загружает устройства из Excel. Возвращает список dict (device_type = имя листа)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    devices = []
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header_row = 0
        for i, row in enumerate(rows):
            cells = [str(c).lower() if c is not None else "" for c in row]
            if any("инв" in c for c in cells):
                header_row = i
                break

        data_rows = rows[header_row + 1 :]

        for row in data_rows:
            if not row or not any(row):
                continue

            def get(idx):
                return row[idx] if idx < len(row) else None

            inv = str(get(2) or "").strip()
            low = inv.lower()
            if low in SKIP_INV or not inv:
                continue
            if inv.isdigit() and len(inv) <= 2:
                continue

            name = str(get(1) or "").strip()
            location = str(get(3) or "").strip()
            last_date = parse_date(get(4))
            expiry_date = parse_date(get(5))
            responsible = str(get(6) or "").strip()

            devices.append(
                {
                    "device_type": sheet_name,
                    "name": name,
                    "inventory_number": inv,
                    "location": location,
                    "last_verification_date": last_date,
                    "expiry_date": expiry_date,
                    "responsible_fio": responsible,
                }
            )
    return devices


def get_existing_device(inv_number: str, device_type: str):
    """Ищет прибор в БД по инв. номеру и типу; expiry_date — последняя по поверкам."""
    conn = get_connection()
    row = conn.execute(
        """
        SELECT d.id, d.type, d.name, d.inventory_number, d.location, d.responsible_fio,
               (SELECT MAX(v.expiry_date) FROM verifications v WHERE v.device_id = d.id) AS expiry_date
        FROM devices d
        WHERE TRIM(IFNULL(d.inventory_number, '')) = TRIM(?)
          AND d.type = ?
        """,
        (inv_number, device_type),
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def insert_device(d: dict) -> None:
    conn = get_connection()
    cur = conn.cursor()
    name = (d.get("name") or "").strip() or "—"
    cur.execute(
        """
        INSERT INTO devices (type, name, inventory_number, location, responsible_fio)
        VALUES (?,?,?,?,?)
        """,
        (
            d["device_type"],
            name,
            d["inventory_number"],
            (d.get("location") or "").strip(),
            (d.get("responsible_fio") or "").strip(),
        ),
    )
    dev_id = cur.lastrowid
    if d.get("expiry_date"):
        cur.execute(
            """
            INSERT INTO verifications (device_id, verification_date, expiry_date, result, comment)
            VALUES (?,?,?,?,?)
            """,
            (
                dev_id,
                d.get("last_verification_date"),
                d["expiry_date"],
                "пройдено",
                "Импорт Excel",
            ),
        )
    conn.commit()
    conn.close()


def update_device(dev_id: int, d: dict) -> None:
    conn = get_connection()
    cur = conn.cursor()
    name = (d.get("name") or "").strip() or "—"
    cur.execute(
        """
        UPDATE devices SET name=?, location=?, responsible_fio=? WHERE id=?
        """,
        (name, (d.get("location") or "").strip(), (d.get("responsible_fio") or "").strip(), dev_id),
    )
    if d.get("expiry_date"):
        cur.execute(
            """
            INSERT INTO verifications (device_id, verification_date, expiry_date, result, comment)
            VALUES (?,?,?,?,?)
            """,
            (
                dev_id,
                d.get("last_verification_date"),
                d["expiry_date"],
                "пройдено",
                "Импорт Excel (обновление)",
            ),
        )
    conn.commit()
    conn.close()


class ImportExcelDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Импорт из Excel")
        self.setMinimumSize(900, 600)
        self.excel_devices = []
        self.conflicts = []
        self.resolutions = {}
        self._new_devices = []
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        file_row = QHBoxLayout()
        self.path_label = QLabel("Файл не выбран")
        self.path_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        btn_choose = QPushButton("📂 Выбрать файл")
        btn_choose.clicked.connect(self._choose_file)
        file_row.addWidget(self.path_label)
        file_row.addWidget(btn_choose)
        layout.addLayout(file_row)

        self.btn_analyze = QPushButton("🔍 Анализировать")
        self.btn_analyze.setEnabled(False)
        self.btn_analyze.clicked.connect(self._analyze)
        layout.addWidget(self.btn_analyze)

        self.status_label = QLabel("")
        layout.addWidget(self.status_label)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        lbl = QLabel("Конфликты (выберите версию для каждой строки):")
        lbl.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        layout.addWidget(lbl)

        self.table = QTableWidget(0, 7)
        self.table.setHorizontalHeaderLabels(
            [
                "Тип",
                "Инв. №",
                "Имя (Excel)",
                "Имя (БД)",
                "Срок (Excel)",
                "Срок (БД)",
                "Решение",
            ]
        )
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.table)

        self.summary_label = QLabel("")
        layout.addWidget(self.summary_label)

        btn_row = QHBoxLayout()
        self.btn_import = QPushButton("✅ Выполнить импорт")
        self.btn_import.setEnabled(False)
        self.btn_import.clicked.connect(self._do_import)
        btn_cancel = QPushButton("Отмена")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(self.btn_import)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

        self.file_path = None

    def _choose_file(self):
        data_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data"
        )
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите Excel файл",
            data_dir if os.path.isdir(data_dir) else "",
            "Excel файлы (*.xlsx *.xls)",
        )
        if path:
            self.file_path = path
            self.path_label.setText(path)
            self.btn_analyze.setEnabled(True)

    def _analyze(self):
        self.status_label.setText("Загрузка файла...")
        self.progress.setVisible(True)
        self.progress.setValue(10)
        try:
            self.excel_devices = load_excel_devices(self.file_path)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось прочитать файл:\n{e}")
            self.progress.setVisible(False)
            return

        self.progress.setValue(40)

        new_devices = []
        self.conflicts = []

        for d in self.excel_devices:
            existing = get_existing_device(d["inventory_number"], d["device_type"])
            if existing is None:
                new_devices.append(d)
                continue

            db_exp = (existing.get("expiry_date") or "").strip()
            ex_exp = (d.get("expiry_date") or "").strip()
            changed = (
                (existing.get("name") or "").strip() != (d.get("name") or "").strip()
                or (existing.get("location") or "").strip() != (d.get("location") or "").strip()
                or (existing.get("responsible_fio") or "").strip()
                != (d.get("responsible_fio") or "").strip()
                or db_exp != ex_exp
            )
            if changed:
                self.conflicts.append((d, existing))

        self.progress.setValue(70)
        self._fill_conflict_table()
        self.progress.setValue(100)
        self.progress.setVisible(False)

        self._new_devices = new_devices
        total = len(self.excel_devices)
        self.status_label.setText(
            f"Найдено в файле: {total}  |  Новых: {len(new_devices)}  |  Конфликтов: {len(self.conflicts)}"
        )
        self.summary_label.setText(
            "Новые будут добавлены автоматически. Для конфликтных строк выберите версию выше."
        )
        self.btn_import.setEnabled(True)

    def _fill_conflict_table(self):
        self.table.setRowCount(0)
        self.resolutions = {}

        for excel_dev, db_dev in self.conflicts:
            key = (excel_dev["inventory_number"], excel_dev["device_type"])
            row = self.table.rowCount()
            self.table.insertRow(row)

            self.table.setItem(row, 0, QTableWidgetItem(excel_dev["device_type"]))
            self.table.setItem(row, 1, QTableWidgetItem(excel_dev["inventory_number"]))

            name_excel = QTableWidgetItem(excel_dev["name"])
            name_db = QTableWidgetItem(db_dev.get("name") or "")
            if excel_dev["name"] != db_dev.get("name", ""):
                name_excel.setBackground(QColor("#fff3cd"))
                name_db.setBackground(QColor("#fff3cd"))
            self.table.setItem(row, 2, name_excel)
            self.table.setItem(row, 3, name_db)

            ex_e = excel_dev.get("expiry_date") or "—"
            db_e = db_dev.get("expiry_date") or "—"
            expiry_excel = QTableWidgetItem(ex_e)
            expiry_db = QTableWidgetItem(db_e)
            if ex_e != db_e:
                expiry_excel.setBackground(QColor("#cfe2ff"))
                expiry_db.setBackground(QColor("#cfe2ff"))
            self.table.setItem(row, 4, expiry_excel)
            self.table.setItem(row, 5, expiry_db)

            combo = QComboBox()
            combo.addItem("Из Excel (новая)", "excel")
            combo.addItem("Из БД (оставить)", "db")
            combo.addItem("Пропустить", "skip")
            combo.setCurrentIndex(0)
            self.resolutions[key] = combo
            self.table.setCellWidget(row, 6, combo)

    def _do_import(self):
        added = 0
        updated = 0
        skipped = 0

        for d in self._new_devices:
            try:
                insert_device(d)
                added += 1
            except Exception as e:
                print(f"Ошибка вставки {d['inventory_number']}: {e}")

        for excel_dev, db_dev in self.conflicts:
            key = (excel_dev["inventory_number"], excel_dev["device_type"])
            combo = self.resolutions.get(key)
            choice = combo.currentData() if combo else "skip"

            if choice == "excel":
                try:
                    update_device(db_dev["id"], excel_dev)
                    updated += 1
                except Exception as e:
                    print(f"Ошибка обновления {excel_dev['inventory_number']}: {e}")
            elif choice == "skip":
                skipped += 1
            else:
                skipped += 1

        QMessageBox.information(
            self,
            "Импорт завершён",
            f"Добавлено: {added}\nОбновлено: {updated}\nПропущено: {skipped}",
        )
        self.accept()
