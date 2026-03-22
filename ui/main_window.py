import os
import sys

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.append(PROJECT_ROOT)

from PyQt6.QtCore import QEvent, Qt, QTimer, QSize, QUrl
from PyQt6.QtGui import (
    QAction,
    QColor,
    QDesktopServices,
    QFont,
    QKeySequence,
    QPainter,
    QPen,
    QShortcut,
)
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QComboBox,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMainWindow,
    QMenu,
    QMessageBox,
    QScrollArea,
    QStyledItemDelegate,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QSizePolicy,
)

from core.journal_dates import journal_reminder_dates
from core.paths import get_help_index_path

from db.database import (
    delete_device,
    get_all_devices,
    get_connection,
    get_location_choices,
    init_database,
    update_device_location,
    update_device_note,
    update_responsible_fio,
)
from core.reset_db import reset_db
from ui.verification_dialog import VerificationDialog
from ui.device_card import DeviceCardDialog
from ui.add_device_dialog import AddDeviceDialog
from ui.date_delegate import DateDelegate
from ui.side_panel import SidePanel
from ui.import_excel_dialog import ImportExcelDialog
from ui.responsible_dialog import ResponsibleDialog
from ui.dashboard_dialog import DashboardDialog
from ui.caldav_settings_dialog import CalDAVSettingsDialog
from ui.email_settings_dialog import EmailSettingsDialog
from ui.word_wrap_delegate import LocationComboDelegate


# ── Делегат: жирная черта под строкой заголовков ──────────────────────────────
class HeaderBorderDelegate(QStyledItemDelegate):
    def paint(self, painter: QPainter, option, index):
        super().paint(painter, option, index)
        if index.row() == 0:
            pen = QPen(QColor("#000000"))
            pen.setWidth(3)
            painter.setPen(pen)
            rect = option.rect
            painter.drawLine(rect.bottomLeft(), rect.bottomRight())


class _UnconstrainedTable(QTableWidget):
    """Снимает искусственный минимум по сумме ширин колонок — окно можно сужать до полос прокрутки."""

    def minimumSizeHint(self):
        return QSize(0, 0)


# ── Константы столбцов (как в data/journal.xlsx, строка заголовков) ─────────
COL_ORD      = 0
COL_LOCATION = 1
COL_NAME     = 2
COL_INV      = 3
COL_TYPE     = 4
COL_EXPIRY   = 5
COL_REM2     = 6
COL_REM1     = 7
COL_REM7     = 8
COL_REM2D    = 9
COL_STATUS   = 10
COL_NOTE     = 11

COLUMNS = [
    "№",
    "Местонахождение (склад/вагон)",
    "Наименование прибора",
    "Инв. номер",
    "Тип прибора",
    "Дата истечения поверки",
    "Дата напоминания\n(-2 месяца)",
    "Дата напоминания\n(-1 месяц)",
    "Дата напоминания\n(-7 дней)",
    "Дата напоминания\n(-2 дня)",
    "Статус",
    "Примечание",
]

COL_FIELDS = {
    COL_LOCATION: "location",
    COL_NAME:     "name",
    COL_INV:      "inventory_number",
    COL_TYPE:     "type",
    COL_EXPIRY:   "expiry_date",
    COL_REM2:     "expiry_date",
    COL_REM1:     "expiry_date",
    COL_REM7:     "expiry_date",
    COL_REM2D:    "expiry_date",
    COL_STATUS:   "status",
    COL_NOTE:     "note",
}

# Отображение статуса в таблице и экспорте — как в journal.xlsx
JOURNAL_STATUS_LABELS = {
    "green":   "🟢 В порядке",
    "yellow":  "🟡 Скоро срок",
    "red":     "🔴 Просрочено",
    "no_data": "⚪ Нет данных",
}

STATUS_ORDER  = {"green": 0, "yellow": 1, "red": 2, "no_data": 3}
STATUS_COLORS = {
    "green":   ("#C6EFCE", "#276221"),
    "yellow":  ("#FFEB9C", "#9C5700"),
    "red":     ("#FFC7CE", "#9C0006"),
    "no_data": ("#D9D9D9", "#444444"),
}
STATUS_LABELS = {
    "green":   "В норме",
    "yellow":  "Скоро срок",
    "red":     "Просрочено",
    "no_data": "Нет данных",
}

HEADER_ROWS = 1
HEADER_TITLE_ROW_HEIGHT = 44
HEADER_SECTION_DRAG_H = 16  # QHeaderView: индексы столбцов + ручной ресайз
TABLE_HEADER_HEIGHT = HEADER_SECTION_DRAG_H + HEADER_TITLE_ROW_HEIGHT + 8
SEARCH_HIGHLIGHT_BG = "#FFF9C4"

# Стартовые ширины колонок (нижняя таблица); ручной ресайз — полоска над table_header
DEFAULT_COLUMN_WIDTHS = [40, 220, 260, 100, 120, 120, 120, 120, 120, 120, 130, 160]
DATA_ROW_HEIGHT = 26


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PoverkiVSE")
        self.resize(1300, 750)
        self.setMinimumSize(0, 0)

        self._sort_col = None
        self._sort_asc = True
        self._search_match_ids = []
        self._search_match_index = -1
        self._highlight_device_id = None
        self._column_widths_seeded = False
        self._column_sync_lock = False

        central = QWidget()
        central.setMinimumSize(0, 0)
        self.setCentralWidget(central)

        # ── главный layout: таблица слева, панель справа ──────────────────
        main_layout = QHBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # ── левая часть ───────────────────────────────────────────────────
        left_widget = QWidget()
        left_widget.setMinimumSize(0, 0)
        left_widget.setSizePolicy(
            QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Expanding
        )
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(6, 6, 6, 6)
        left_layout.setSpacing(6)

        # фильтры — в скролле: иначе сумма minWidth комбо задаёт огромный minimumSize окна
        filters_host = QWidget()
        filters_host.setMinimumSize(0, 0)
        filters_layout = QHBoxLayout(filters_host)
        filters_layout.setContentsMargins(0, 0, 0, 0)

        filters_layout.addWidget(QLabel("Тип:"))
        self.type_filter = QComboBox()
        self.type_filter.addItems(["Все", "Алкометры", "Тонометры"])
        self.type_filter.currentIndexChanged.connect(self.refresh_table)
        filters_layout.addWidget(self.type_filter)

        filters_layout.addWidget(QLabel("Статус:"))
        self.status_filter = QComboBox()
        self.status_filter.addItems(["Все", "🟢 Зелёный", "🟡 Жёлтый", "🔴 Красный", "⚪ Нет данных"])
        self.status_filter.currentIndexChanged.connect(self.refresh_table)
        filters_layout.addWidget(self.status_filter)

        filters_layout.addWidget(QLabel("Местонахождение (склад/вагон):"))
        self.location_filter = QComboBox()
        self.location_filter.addItem("Все")
        self.location_filter.currentIndexChanged.connect(self.refresh_table)
        filters_layout.addWidget(self.location_filter)

        filters_layout.addWidget(QLabel("Ответственный:"))
        self.responsible_filter = QComboBox()
        self.responsible_filter.addItem("Все")
        self.responsible_filter.currentIndexChanged.connect(self.refresh_table)
        filters_layout.addWidget(self.responsible_filter)

        filters_layout.addStretch()
        for cb in (
            self.type_filter,
            self.status_filter,
            self.location_filter,
            self.responsible_filter,
        ):
            cb.setMinimumWidth(0)
            cb.setSizePolicy(
                QSizePolicy.Policy.MinimumExpanding, QSizePolicy.Policy.Fixed
            )

        filters_scroll = QScrollArea()
        filters_scroll.setWidgetResizable(True)
        filters_scroll.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAsNeeded
        )
        filters_scroll.setVerticalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOff
        )
        filters_scroll.setFrameShape(QFrame.Shape.NoFrame)
        filters_scroll.setMinimumHeight(32)
        filters_scroll.setMaximumHeight(44)
        filters_scroll.setWidget(filters_host)
        filters_scroll.setMinimumSize(0, 0)
        self.filters_scroll = filters_scroll

        # Верхняя таблица — только 2 строки заголовка (не прокручивается по вертикали)
        self.table_header = _UnconstrainedTable()
        self._header_delegate = HeaderBorderDelegate()
        self.table_header.setItemDelegate(self._header_delegate)
        self.table_header.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table_header.horizontalHeader().setVisible(False)
        self.table_header.verticalHeader().setVisible(False)
        self.table_header.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.table_header.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.table_header.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_header.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        self.table_header.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        hh = self.table_header.horizontalHeader()
        hh.setVisible(True)
        hh.setFixedHeight(HEADER_SECTION_DRAG_H)
        hh.setHighlightSections(False)
        hh.setStyleSheet(
            "QHeaderView::section { background: #D9D9D9; color: #555555; "
            "font: 8pt Calibri; border: 1px solid #b0b0b0; padding: 2px; }"
        )
        for c in range(len(COLUMNS)):
            hh.setSectionResizeMode(c, QHeaderView.ResizeMode.Interactive)
        self.table_header.cellClicked.connect(self.on_header_cell_clicked)
        self.table_header.setFrameShape(QFrame.Shape.NoFrame)
        self.table_header.setFixedHeight(TABLE_HEADER_HEIGHT)
        self.table_header.setSizePolicy(
            QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Fixed
        )

        # Нижняя таблица — только строки данных
        self.table = _UnconstrainedTable()
        self.table.setMinimumSize(0, 0)
        self.table.setSizePolicy(
            QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Fixed
        )
        self._date_delegate = DateDelegate()
        self._location_delegate = LocationComboDelegate(self.table)
        self.table.setItemDelegateForColumn(COL_EXPIRY, self._date_delegate)
        self.table.setItemDelegateForColumn(COL_LOCATION, self._location_delegate)
        self.table.setEditTriggers(
            QTableWidget.EditTrigger.DoubleClicked
            | QTableWidget.EditTrigger.SelectedClicked
            | QTableWidget.EditTrigger.EditKeyPressed
        )
        # Ресайз столбцов — полоска над table_header; здесь заголовок скрыт
        dh = self.table.horizontalHeader()
        dh.setVisible(False)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(DATA_ROW_HEIGHT)
        self.table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._on_table_context_menu)
        self._shortcut_delete = QShortcut(QKeySequence(Qt.Key.Key_Delete), self.table)
        self._shortcut_delete.activated.connect(self._delete_selected_devices)
        self.table.setAlternatingRowColors(False)
        self.table.cellDoubleClicked.connect(self.on_cell_double_clicked)
        self.table.itemChanged.connect(self.on_item_changed)
        dh.setMinimumSectionSize(24)
        for c in range(len(COLUMNS)):
            dh.setSectionResizeMode(c, QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().sectionResized.connect(self._on_data_column_resized)
        self.table_header.horizontalHeader().sectionResized.connect(
            self._on_header_column_resized
        )

        # Левая колонка: фильтры + заголовок таблицы + данные — одна прокрутка (вертикаль и горизонталь)
        self._left_scroll = QScrollArea()
        self._left_scroll.setFrameShape(QFrame.Shape.NoFrame)
        self._left_scroll.setWidgetResizable(False)
        self._left_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self._left_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self._left_content = QWidget()
        self._left_content.setMinimumSize(0, 0)
        _left_v = QVBoxLayout(self._left_content)
        _left_v.setContentsMargins(0, 0, 0, 0)
        _left_v.setSpacing(6)
        _left_v.addWidget(self.filters_scroll)
        _left_v.addWidget(self.table_header)
        _left_v.addWidget(self.table)
        self._left_scroll.setWidget(self._left_content)
        self._left_scroll.viewport().installEventFilter(self)

        left_layout.addWidget(self._left_scroll, stretch=1)

        main_layout.addWidget(left_widget, stretch=1)

        # ── правая панель ─────────────────────────────────────────────────
        self.side_panel = SidePanel(self)
        self.side_panel.sig_add_device.connect(self.on_add_device)
        self.side_panel.sig_notify.connect(self.on_send_notifications)
        self.side_panel.sig_export.connect(self.on_export_excel)
        self.side_panel.sig_import_excel.connect(self.on_import_excel)
        self.side_panel.sig_responsible.connect(self.on_responsible)
        self.side_panel.sig_dashboard.connect(self.on_dashboard)
        self.side_panel.sig_caldav_yandex.connect(self.on_caldav_yandex)
        self.side_panel.sig_caldav_mailru.connect(self.on_caldav_mailru)
        self.side_panel.sig_email_settings.connect(self.on_email_settings)
        self.side_panel.sig_clear_database.connect(self.on_clear_database)
        self.side_panel.sig_help.connect(self.on_open_help)
        self.side_panel.search_box.textChanged.connect(self._on_search_text_changed)
        self.side_panel.search_box.returnPressed.connect(self._search_nav_find_first)
        self.side_panel.sig_search_find.connect(self._search_nav_find_first)
        self.side_panel.sig_search_next.connect(self._search_nav_next)
        self.side_panel.sig_search_prev.connect(self._search_nav_prev)
        main_layout.addWidget(self.side_panel)

        init_database()
        self.refresh_table()
        self.fill_filter_values()
        self._update_search_nav_buttons()

    # ── вспомогательные ───────────────────────────────────────────────────
    def _status_from_filter(self):
        return {
            "🟢 Зелёный":   "green",
            "🟡 Жёлтый":    "yellow",
            "🔴 Красный":   "red",
            "⚪ Нет данных": "no_data",
        }.get(self.status_filter.currentText(), "")

    def _status_key_from_label(self, label: str) -> str:
        t = (label or "").strip()
        for key, jl in JOURNAL_STATUS_LABELS.items():
            if jl == t:
                return key
        return {
            "В норме":    "green",
            "Скоро срок": "yellow",
            "Просрочено": "red",
        }.get(t, "no_data")

    def _device_id_from_row(self, row: int) -> int | None:
        it = self.table.item(row, COL_ORD)
        if not it:
            return None
        v = it.data(Qt.ItemDataRole.UserRole)
        if v is None:
            return None
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    def _device_ids_from_selection(self) -> list[int]:
        rows = {idx.row() for idx in self.table.selectedIndexes()}
        ids: list[int] = []
        for r in sorted(rows):
            did = self._device_id_from_row(r)
            if did is not None:
                ids.append(did)
        return ids

    def _on_table_context_menu(self, pos):
        ids = self._device_ids_from_selection()
        menu = QMenu(self.table)
        if len(ids) == 1:
            act = QAction("Удалить прибор…", self.table)
        elif len(ids) > 1:
            act = QAction(f"Удалить выбранные ({len(ids)})…", self.table)
        else:
            act = QAction("Удалить выбранные…", self.table)
            act.setEnabled(False)
        act.triggered.connect(self._delete_selected_devices)
        menu.addAction(act)
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def _delete_selected_devices(self):
        ids = self._device_ids_from_selection()
        if not ids:
            QMessageBox.information(
                self,
                "Удаление",
                "Выберите одну или несколько строк в таблице.",
            )
            return
        n = len(ids)
        if n == 1:
            msg = "Удалить этот прибор из базы? Связанные поверки и документы будут удалены."
        else:
            msg = (
                f"Удалить {n} приборов из базы? Связанные поверки и документы будут удалены."
            )
        if (
            QMessageBox.question(
                self,
                "Удаление",
                msg,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            != QMessageBox.StandardButton.Yes
        ):
            return
        for did in ids:
            delete_device(did)
        self.refresh_table()
        self.fill_filter_values()
        self._rebuild_search_matches()
        self._update_search_nav_buttons()

    def _device_full_by_id(self, device_id: int):
        for d in get_all_devices():
            if d.get("id") == device_id:
                return d
        return None

    def _row_matches_full_db_search(self, row: dict, q: str) -> bool:
        """Подстрока в любом из отображаемых полей прибора (без учёта регистра)."""
        if not q:
            return False
        st = row.get("status") or "no_data"
        parts = [
            str(row.get("id") or ""),
            row.get("type") or "",
            row.get("inventory_number") or "",
            row.get("serial_number") or "",
            row.get("name") or "",
            row.get("location") or "",
            row.get("note") or "",
            row.get("responsible_fio") or "",
            row.get("expiry_date") or "",
            row.get("verification_date") or "",
            JOURNAL_STATUS_LABELS.get(st, ""),
            STATUS_LABELS.get(st, ""),
        ]
        for s in parts:
            if q in (s or "").lower():
                return True
        return False

    def _rebuild_search_matches(self):
        q = self.side_panel.search_box.text().strip().lower()
        self._search_match_ids = []
        if not q:
            return
        for row in get_all_devices():
            if self._row_matches_full_db_search(row, q):
                self._search_match_ids.append(row["id"])

    def _on_search_text_changed(self, _text=None):
        self._rebuild_search_matches()
        self._search_match_index = -1
        self._highlight_device_id = None
        self._update_search_nav_buttons()
        self.refresh_table()

    def _update_search_nav_buttons(self):
        ok = bool(self.side_panel.search_box.text().strip()) and bool(self._search_match_ids)
        self.side_panel.search_btn_find.setEnabled(ok)
        self.side_panel.search_btn_next.setEnabled(ok)
        self.side_panel.search_btn_prev.setEnabled(ok)

    def _reset_filters_to_all(self):
        self.type_filter.setCurrentIndex(0)
        self.status_filter.setCurrentIndex(0)
        self.location_filter.setCurrentIndex(0)
        self.responsible_filter.setCurrentIndex(0)

    def _collect_filtered_devices(self):
        """Строки таблицы с учётом комбобоксов (без текста поиска)."""
        all_devices = get_all_devices()
        type_filter = self.type_filter.currentText()
        status_filter = self._status_from_filter()
        loc_filter = self.location_filter.currentText()
        resp_filter = self.responsible_filter.currentText()
        filtered = []
        for row in all_devices:
            if type_filter != "Все" and row.get("type") != type_filter:
                continue
            if status_filter and row.get("status") != status_filter:
                continue
            if loc_filter != "Все" and row.get("location") != loc_filter:
                continue
            if resp_filter != "Все" and row.get("responsible_fio") != resp_filter:
                continue
            filtered.append(row)
        return filtered

    def _scroll_to_device_id(self, dev_id: int):
        for r in range(self.table.rowCount()):
            rid = self._device_id_from_row(r)
            if rid is not None and int(rid) == int(dev_id):
                it = self.table.item(r, COL_ORD)
                if it:
                    self.table.selectRow(r)
                    self.table.scrollToItem(it)
                return

    def _go_to_current_search_match(self):
        if not self._search_match_ids or self._search_match_index < 0:
            return
        dev_id = self._search_match_ids[self._search_match_index]
        self._highlight_device_id = dev_id
        vis = {r["id"] for r in self._collect_filtered_devices()}
        if dev_id not in vis:
            self._reset_filters_to_all()
        self.refresh_table()
        self.fill_filter_values()
        self._scroll_to_device_id(dev_id)
        self._update_search_nav_buttons()

    def _search_nav_find_first(self):
        self._rebuild_search_matches()
        if not self._search_match_ids:
            self._highlight_device_id = None
            self.refresh_table()
            self._update_search_nav_buttons()
            return
        self._search_match_index = 0
        self._go_to_current_search_match()

    def _search_nav_next(self):
        self._rebuild_search_matches()
        if not self._search_match_ids:
            return
        if self._search_match_index < 0:
            self._search_match_index = 0
        else:
            self._search_match_index = (self._search_match_index + 1) % len(
                self._search_match_ids
            )
        self._go_to_current_search_match()

    def _search_nav_prev(self):
        self._rebuild_search_matches()
        if not self._search_match_ids:
            return
        if self._search_match_index <= 0:
            self._search_match_index = len(self._search_match_ids) - 1
        else:
            self._search_match_index -= 1
        self._go_to_current_search_match()

    def _table_cell_text(self, row: int, col: int) -> str:
        it = self.table.item(row, col)
        return it.text() if it else ""

    def _on_data_column_resized(self, logical_index: int, old_size: int, new_size: int):
        if self._column_sync_lock:
            return
        self._column_sync_lock = True
        self.table_header.setColumnWidth(logical_index, new_size)
        self._column_sync_lock = False
        QTimer.singleShot(0, self._update_left_scroll_geometry)

    def _on_header_column_resized(self, logical_index: int, old_size: int, new_size: int):
        if self._column_sync_lock:
            return
        self._column_sync_lock = True
        self.table.setColumnWidth(logical_index, new_size)
        self._column_sync_lock = False
        QTimer.singleShot(0, self._update_left_scroll_geometry)

    def _copy_column_widths_to_header(self):
        for c in range(len(COLUMNS)):
            self.table_header.setColumnWidth(c, self.table.columnWidth(c))

    def _update_left_scroll_geometry(self) -> None:
        """Размер содержимого левой колонки: фильтры + заголовок + все строки; прокрутка у области _left_scroll."""
        if not hasattr(self, "_left_scroll"):
            return
        vp = self._left_scroll.viewport()
        if not vp or vp.width() < 1:
            return
        vw = max(vp.width(), 1)
        tw = sum(max(self.table.columnWidth(c), 0) for c in range(len(COLUMNS)))
        inner_w = max(tw, vw)
        n = self.table.rowCount()
        body_h = max(n * DATA_ROW_HEIGHT, 50)
        fh = max(
            int(self.filters_scroll.sizeHint().height()),
            int(self.filters_scroll.minimumHeight()),
        )
        inner_h = fh + TABLE_HEADER_HEIGHT + body_h
        self._left_content.setFixedSize(inner_w, inner_h)
        self.filters_scroll.setFixedWidth(inner_w)
        self.table_header.setFixedWidth(inner_w)
        self.table.setFixedWidth(inner_w)
        self.table.setFixedHeight(body_h)

    def eventFilter(self, obj, event):
        if (
            hasattr(self, "_left_scroll")
            and obj is self._left_scroll.viewport()
            and event.type() == QEvent.Type.Resize
        ):
            QTimer.singleShot(0, self._update_left_scroll_geometry)
        return super().eventFilter(obj, event)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        QTimer.singleShot(0, self._update_left_scroll_geometry)

    # ── заполнение фильтров ───────────────────────────────────────────────
    def fill_filter_values(self):
        devices = get_all_devices()
        locations    = sorted({r["location"] for r in devices if r.get("location")})
        responsibles = sorted({r["responsible_fio"] for r in devices if r.get("responsible_fio")})

        cur_loc  = self.location_filter.currentText()
        cur_resp = self.responsible_filter.currentText()

        self.location_filter.blockSignals(True)
        self.responsible_filter.blockSignals(True)

        self.location_filter.clear()
        self.location_filter.addItem("Все")
        self.location_filter.addItems(locations)

        self.responsible_filter.clear()
        self.responsible_filter.addItem("Все")
        self.responsible_filter.addItems(responsibles)

        idx = self.location_filter.findText(cur_loc)
        if idx != -1:
            self.location_filter.setCurrentIndex(idx)

        idx = self.responsible_filter.findText(cur_resp)
        if idx != -1:
            self.responsible_filter.setCurrentIndex(idx)

        self.location_filter.blockSignals(False)
        self.responsible_filter.blockSignals(False)

    # ── основная отрисовка таблицы ────────────────────────────────────────
    def refresh_table(self):
        self.table.blockSignals(True)
        self.table_header.blockSignals(True)

        filtered = self._collect_filtered_devices()
        base_locs = get_location_choices()

        if self._sort_col is not None:
            field = COL_FIELDS.get(self._sort_col)
            if field:
                if field == "status":
                    filtered.sort(
                        key=lambda r: STATUS_ORDER.get(r.get("status") or "no_data", 9),
                        reverse=not self._sort_asc,
                    )
                else:
                    filtered.sort(
                        key=lambda r: (r.get(field) or ""),
                        reverse=not self._sort_asc,
                    )

        # ── верхняя таблица: индексы 0…n в QHeaderView (одна строка), ниже — названия столбцов
        self.table_header.clearContents()
        self.table_header.setRowCount(HEADER_ROWS)
        self.table_header.setColumnCount(len(COLUMNS))
        self.table_header.setHorizontalHeaderLabels([str(c) for c in range(len(COLUMNS))])

        hdr_font = QFont("Calibri", 10)
        hdr_font.setBold(True)
        for c, col_name in enumerate(COLUMNS):
            arrow = ""
            if self._sort_col == c:
                arrow = " ▲" if self._sort_asc else " ▼"
            cell = QTableWidgetItem(col_name + arrow)
            cell.setFont(hdr_font)
            cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            cell.setFlags(Qt.ItemFlag.ItemIsEnabled)
            cell.setBackground(QColor("#1F3864"))
            cell.setForeground(QColor("#FFFFFF"))
            self.table_header.setItem(0, c, cell)
        self.table_header.setRowHeight(0, HEADER_TITLE_ROW_HEIGHT)

        # ── нижняя таблица: только данные ──
        self.table.clearContents()
        self.table.setRowCount(len(filtered))
        self.table.setColumnCount(len(COLUMNS))
        self.table.setHorizontalHeaderLabels([""] * len(COLUMNS))

        if not self._column_widths_seeded:
            for c, w in enumerate(DEFAULT_COLUMN_WIDTHS):
                self.table.setColumnWidth(c, w)
            self._column_widths_seeded = True

        for r, row in enumerate(filtered):
            status = row.get("status") or "no_data"
            expiry = row.get("expiry_date") or ""
            dev_id = row.get("id")
            is_hl = (
                self._highlight_device_id is not None
                and dev_id is not None
                and int(dev_id) == int(self._highlight_device_id)
            )

            rem = journal_reminder_dates(expiry)
            values = [
                str(r + 1),
                row.get("location") or "",
                row.get("name") or "",
                row.get("inventory_number") or "",
                row.get("type") or "",
                expiry,
                rem[0],
                rem[1],
                rem[2],
                rem[3],
                JOURNAL_STATUS_LABELS.get(status, "⚪ Нет данных"),
                row.get("note") or "",
            ]

            bg_status, fg_status = STATUS_COLORS.get(status, STATUS_COLORS["no_data"])

            cur_loc = (row.get("location") or "").strip()
            loc_items = list(base_locs)
            if cur_loc and cur_loc not in loc_items:
                loc_items.append(cur_loc)
                loc_items.sort()

            for c, val in enumerate(values):
                if c == COL_ORD:
                    ord_cell = QTableWidgetItem(val)
                    ord_cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    ord_cell.setData(Qt.ItemDataRole.UserRole, dev_id)
                    ord_cell.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                    if is_hl:
                        ord_cell.setBackground(QColor(SEARCH_HIGHLIGHT_BG))
                    else:
                        ord_cell.setBackground(QColor("#FFFFFF"))
                    ord_cell.setForeground(QColor("#000000"))
                    self.table.setItem(r, COL_ORD, ord_cell)
                    continue
                if c == COL_LOCATION:
                    loc_cell = QTableWidgetItem(cur_loc)
                    loc_cell.setTextAlignment(
                        Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
                    )
                    loc_cell.setData(Qt.ItemDataRole.UserRole, loc_items)
                    loc_cell.setFlags(
                        Qt.ItemFlag.ItemIsEnabled
                        | Qt.ItemFlag.ItemIsSelectable
                        | Qt.ItemFlag.ItemIsEditable
                    )
                    if is_hl:
                        loc_cell.setBackground(QColor(SEARCH_HIGHLIGHT_BG))
                    else:
                        loc_cell.setBackground(QColor("#FFFFFF"))
                    loc_cell.setForeground(QColor("#000000"))
                    self.table.setItem(r, COL_LOCATION, loc_cell)
                    continue
                cell = QTableWidgetItem("" if val is None else str(val))
                if c in (COL_NAME, COL_NOTE):
                    cell.setTextAlignment(
                        Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
                    )
                else:
                    cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                if c == COL_STATUS:
                    cell.setBackground(QColor(bg_status))
                    cell.setForeground(QColor(fg_status))
                    cell.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                elif c in (COL_REM2, COL_REM1, COL_REM7, COL_REM2D, COL_NAME, COL_INV, COL_TYPE):
                    cell.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                    if is_hl:
                        cell.setBackground(QColor(SEARCH_HIGHLIGHT_BG))
                    else:
                        cell.setBackground(QColor("#FFFFFF"))
                    cell.setForeground(QColor("#000000"))
                elif c == COL_EXPIRY:
                    cell.setBackground(QColor("#FFFFFF"))
                    cell.setForeground(QColor("#000000"))
                    cell.setFlags(
                        Qt.ItemFlag.ItemIsEnabled
                        | Qt.ItemFlag.ItemIsSelectable
                        | Qt.ItemFlag.ItemIsEditable
                    )
                elif c == COL_NOTE:
                    cell.setBackground(QColor("#FFFFFF"))
                    cell.setForeground(QColor("#000000"))
                    cell.setFlags(
                        Qt.ItemFlag.ItemIsEnabled
                        | Qt.ItemFlag.ItemIsSelectable
                        | Qt.ItemFlag.ItemIsEditable
                    )
                else:
                    if is_hl:
                        cell.setBackground(QColor(SEARCH_HIGHLIGHT_BG))
                    else:
                        cell.setBackground(QColor("#FFFFFF"))
                    cell.setForeground(QColor("#000000"))
                    cell.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)

                self.table.setItem(r, c, cell)

        self._copy_column_widths_to_header()
        for rr in range(self.table.rowCount()):
            self.table.setRowHeight(rr, DATA_ROW_HEIGHT)

        total = len(filtered)
        green = sum(1 for r in filtered if r.get("status") == "green")
        yellow = sum(1 for r in filtered if r.get("status") == "yellow")
        red = sum(1 for r in filtered if r.get("status") == "red")
        self.side_panel.update_counter(total, green, yellow, red)

        self.table_header.blockSignals(False)
        self.table.blockSignals(False)

        QTimer.singleShot(0, self._update_left_scroll_geometry)

    # ── клик по строке заголовков (верхняя таблица) ─────────────────────────
    def on_header_cell_clicked(self, row, col):
        if row == 0:
            if col == COL_ORD:
                return
            if self._sort_col == col:
                self._sort_asc = not self._sort_asc
            else:
                self._sort_col = col
                self._sort_asc = True
            self.refresh_table()

    def on_cell_double_clicked(self, row, col):
        if col in (
            COL_ORD,
            COL_LOCATION,
            COL_EXPIRY,
            COL_REM2,
            COL_REM1,
            COL_REM7,
            COL_REM2D,
        ):
            return

        device_id = self._device_id_from_row(row)
        if device_id is None:
            return

        def cell_text(c):
            return self._table_cell_text(row, c)

        device_data = {
            "id":               device_id,
            "type":             cell_text(COL_TYPE),
            "inventory_number": cell_text(COL_INV),
            "location":         cell_text(COL_LOCATION),
            "expiry_date":      cell_text(COL_EXPIRY),
            "responsible_fio":  "",
            "status":           self._status_key_from_label(cell_text(COL_STATUS)),
            "verification_date": "",
            "name":             cell_text(COL_NAME),
        }
        full = self._device_full_by_id(device_id)
        if full:
            device_data["verification_date"] = full.get("verification_date") or ""
            device_data["serial_number"] = full.get("serial_number") or ""
            device_data["name"] = full.get("name") or device_data["name"]
            device_data["responsible_fio"] = full.get("responsible_fio") or ""
        else:
            device_data["serial_number"] = ""

        dlg = DeviceCardDialog(device_data, self)
        if dlg.exec():
            init_database()
            self.refresh_table()
            self.fill_filter_values()

    def on_item_changed(self, item):
        row = item.row()
        col = item.column()

        device_id = self._device_id_from_row(row)
        if device_id is None:
            return

        if col == COL_EXPIRY:
            new_date = item.text().strip()
            parts = new_date.split("-")
            if len(parts) != 3:
                return
            conn = get_connection()
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO verifications (device_id, verification_date, expiry_date, result, comment)
                VALUES (?, ?, ?, ?, ?)
                """,
                (device_id, None, new_date, "пройдено", "Ручное редактирование"),
            )
            conn.commit()
            conn.close()
            init_database()
            self.refresh_table()
            self.fill_filter_values()

        elif col == COL_NOTE:
            update_device_note(device_id, item.text())

        elif col == COL_LOCATION:
            new_loc = item.text().strip()
            full = self._device_full_by_id(device_id)
            old_loc = (full or {}).get("location") or ""
            if new_loc != old_loc:
                update_device_location(device_id, new_loc)
                self.fill_filter_values()

    # ── обработчики кнопок из правой панели ───────────────────────────────
    def on_add_device(self):
        dlg = AddDeviceDialog(self)
        if dlg.exec():
            init_database()
            self.refresh_table()
            self.fill_filter_values()
            last_row = self.table.rowCount() - 1
            if last_row >= 0:
                self.table.scrollToItem(self.table.item(last_row, 0))
                self.table.selectRow(last_row)

    def on_import_excel(self):
        dlg = ImportExcelDialog(self)
        if dlg.exec():
            init_database()
            self.refresh_table()
            self.fill_filter_values()

    def on_responsible(self):
        ResponsibleDialog(self).exec()

    def on_dashboard(self):
        DashboardDialog(self).exec()

    def on_caldav_yandex(self):
        CalDAVSettingsDialog(provider="yandex", parent=self).exec()

    def on_caldav_mailru(self):
        CalDAVSettingsDialog(provider="mailru", parent=self).exec()

    def on_email_settings(self):
        EmailSettingsDialog(self).exec()

    def on_open_help(self):
        path = get_help_index_path()
        if not os.path.isfile(path):
            QMessageBox.warning(
                self,
                "Справка",
                "Файл справки не найден.\n"
                f"Ожидался:\n{path}",
            )
            return
        url = QUrl.fromLocalFile(os.path.normpath(path))
        if not QDesktopServices.openUrl(url):
            QMessageBox.warning(
                self,
                "Справка",
                "Не удалось открыть справку в браузере.",
            )

    def on_clear_database(self):
        reply = QMessageBox.question(
            self,
            "Очистка базы данных",
            "Будут удалены все приборы, история поверок, лог уведомлений и файлы в папке documents.\n"
            "Настройки (CalDAV, email) и справочники пользователей сохранятся.\n\n"
            "Продолжить?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        second = QMessageBox.warning(
            self,
            "Подтверждение",
            "Это действие нельзя отменить. Удалить все данные приборов?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if second != QMessageBox.StandardButton.Yes:
            return
        reset_db()
        init_database()
        self.side_panel.search_box.clear()
        self._rebuild_search_matches()
        self._search_match_index = -1
        self._highlight_device_id = None
        self.refresh_table()
        self.fill_filter_values()
        self._update_search_nav_buttons()
        QMessageBox.information(self, "Готово", "База данных очищена.")

    def on_send_notifications(self):
        reply = QMessageBox.question(
            self,
            "Отправить уведомления",
            "Отправить в MAX утреннюю сводную ведомость\n"
            "(стадии, сроки, ответственные — без списка каждого прибора)?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        self.side_panel.notify_btn.setEnabled(False)
        self.side_panel.notify_btn.setText("⏳  Отправка...")

        try:
            from core.notifier import check_and_notify
            result = check_and_notify(dry_run=False)
            dup = result.get("skipped_already_today", 0)
            QMessageBox.information(
                self,
                "Готово",
                f"✅ Отправлено сообщений: {result['sent']}\n"
                f"⏭  Уже сегодня (не дублировали): {dup}\n"
                f"❌ Ошибок: {result['errors']}",
            )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))
        finally:
            self.side_panel.notify_btn.setEnabled(True)
            self.side_panel.notify_btn.setText("🔔  Уведомления")

    def on_export_excel(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить журнал поверок",
            f"Журнал_поверок_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Журнал поверок"

        headers = list(COLUMNS)
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1F3864")

        STATUS_COLORS_XL = {
            "green":   "C6EFCE",
            "yellow":  "FFEB9C",
            "red":     "FFC7CE",
            "no_data": "D9D9D9",
        }

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = border

        ws.row_dimensions[1].height = 44

        devices = get_all_devices()
        type_f   = self.type_filter.currentText()
        status_f = self._status_from_filter()
        loc_f    = self.location_filter.currentText()
        resp_f   = self.responsible_filter.currentText()

        filtered = []
        for row in devices:
            if type_f != "Все" and row.get("type") != type_f:
                continue
            if status_f and row.get("status") != status_f:
                continue
            if loc_f != "Все" and row.get("location") != loc_f:
                continue
            if resp_f != "Все" and row.get("responsible_fio") != resp_f:
                continue
            filtered.append(row)

        for excel_row, (ord_i, row) in enumerate(
            ((i + 1, r) for i, r in enumerate(filtered)), start=2
        ):
            status = row.get("status") or "no_data"
            exp = row.get("expiry_date") or ""
            rem = journal_reminder_dates(exp)
            values = [
                ord_i,
                row.get("location") or "",
                row.get("name") or "",
                row.get("inventory_number") or "",
                row.get("type") or "",
                exp,
                rem[0],
                rem[1],
                rem[2],
                rem[3],
                JOURNAL_STATUS_LABELS.get(status, "⚪ Нет данных"),
                row.get("note") or "",
            ]
            bg = STATUS_COLORS_XL.get(status, "FFFFFF")
            fill = PatternFill("solid", fgColor=bg)

            for c, val in enumerate(values, 1):
                cell = ws.cell(row=excel_row, column=c, value=val)
                cell.fill = fill
                cell.border = border
                h = "center"
                if c in (2, 3, 12):
                    h = "left"
                cell.alignment = Alignment(horizontal=h, vertical="center", wrap_text=False)
            ws.row_dimensions[excel_row].height = 18

        col_widths = [6, 36, 40, 14, 16, 16, 16, 16, 16, 16, 18, 28]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        wb.save(path)
        QMessageBox.information(
            self, "Экспорт завершён",
            f"Файл сохранён:\n{path}\n\nСтрок экспортировано: {len(filtered)}"
        )
